import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator, AutoMinorLocator
from sklearn.neighbors import LocalOutlierFactor
from sklearn.preprocessing import StandardScaler
from scipy.stats import ks_2samp, skew, kurtosis, norm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from mpl_toolkits.mplot3d import Axes3D
import subprocess
import matplotlib.font_manager as fm

# ============================== 字体设置 ==============================
font_path = "C:/Windows/Fonts/times.ttf"
times_new_roman = fm.FontProperties(fname=font_path).get_name()
plt.rcParams['font.family'] = times_new_roman
plt.rcParams['axes.unicode_minus'] = False
label_font = fm.FontProperties(fname=font_path, weight='bold')

# ============================== 图像保存函数 ==============================
INKSCAPE_PATH = r"C:\Program Files\Inkscape\bin\inkscape.com"
def save_figure(name):
    plt.savefig(f"{name}.svg", bbox_inches='tight')
    plt.savefig(f"{name}.png", dpi=300, bbox_inches='tight')
    try:
        subprocess.run([INKSCAPE_PATH, f"{name}.svg", "--export-type=emf", "--export-filename", f"{name}.emf"], check=True)
        print(f"✅ Saved: {name}.svg, {name}.png, {name}.emf")
    except subprocess.CalledProcessError as e:
        print(f"❌ Inkscape export failed for {name}: {e}")

# ============================== 数据加载与清洗 ==============================
cols_mapping = {'推进速度': 'Tunnelling Speed', '扭矩': 'Torque', '刀盘推力': 'Cutterhead Thrust'}
df = pd.read_excel('11.xlsx')
df_eng = df.rename(columns=cols_mapping)
cols = list(cols_mapping.values())

scaler = StandardScaler()
df_scaled = scaler.fit_transform(df_eng[cols])
lof = LocalOutlierFactor(n_neighbors=5, contamination=0.02)
outliers = lof.fit_predict(df_scaled) == -1

df_clean = df_eng.copy()
df_clean.loc[outliers, cols] = np.nan
df_clean[cols] = df_clean[cols].interpolate(method='linear')

with pd.ExcelWriter('11_clean.xlsx', engine='openpyxl', mode='w') as writer:
    df_clean.to_excel(writer, sheet_name='Cleaned Data', index=False)

wb = load_workbook('11_clean.xlsx')
ws = wb['Cleaned Data']
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
for row_idx in df.index[outliers] + 2:
    for col_idx in range(1, len(cols) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = red_fill
wb.save('11_clean.xlsx')

# ============================== 评估函数 ==============================
def evaluate_cleaning(original, cleaned, outliers):
    stats_df = pd.DataFrame({
        'Original Mean': original.mean(),
        'Cleaned Mean': cleaned.mean(),
        'Original Variance': original.var(),
        'Cleaned Variance': cleaned.var(),
        'Original Skewness': original.apply(skew),
        'Cleaned Skewness': cleaned.apply(skew),
        'Original Kurtosis': original.apply(kurtosis),
        'Cleaned Kurtosis': cleaned.apply(kurtosis),
        'Outlier Ratio': outliers.mean()
    })
    ks_df = pd.DataFrame({col: ks_2samp(original[col].dropna(), cleaned[col].dropna()).pvalue for col in original.columns}, index=['KS Test p-value']).T
    return stats_df, ks_df

stats_df, ks_df = evaluate_cleaning(df_eng[cols], df_clean[cols], outliers)
print("\n=== Statistical Evaluation ===")
print(stats_df)
print("\n=== KS Test Results ===")
print(ks_df)

# ============================== 图表绘制函数 ==============================
def plot_boxplot(cleaned_data):
    plt.figure(figsize=(10, 6))
    cleaned_data.boxplot()
    plt.title("Outlier Distribution Boxplot", fontsize=12)
    plt.ylabel("Value", fontsize=12, fontweight='bold')
    plt.xticks(rotation=45, fontweight='bold')
    plt.yticks(fontweight='bold')
    plt.tight_layout()
    save_figure('outlier_boxplot')
    plt.close()

def plot_comparison(original, cleaned, parameters):
    fig, axes = plt.subplots(len(parameters), 1, figsize=(14, 12))
    for i, param in enumerate(parameters):
        axes[i].yaxis.set_major_locator(MaxNLocator(nbins=10))
        axes[i].yaxis.set_minor_locator(AutoMinorLocator(5))
        axes[i].tick_params(axis='y', which='minor', length=3, direction='in')
        axes[i].tick_params(axis='both', which='major', direction='in', labelsize=18, width=2, length=6, color='black', labelcolor='black')
        axes[i].plot(original.index, original[param], color='#444444', linewidth=2.5, label='Original Data')
        axes[i].plot(cleaned[param], color='#E41A1C', linewidth=3.0, linestyle='--', label='Cleaned Data')
        if outliers.any():
            axes[i].scatter(original.index[outliers], original[param].iloc[outliers], color='#4DAF4A', edgecolor='black', s=50, linewidths=1.5, zorder=3, label='Outliers')
        axes[i].legend(loc='lower right', frameon=True, shadow=True, fontsize=14, borderpad=0.8, edgecolor='black')
        axes[i].set_title(f"{param}", fontsize=22, fontweight='bold')
        axes[i].set_ylabel("Value", fontsize=24, fontweight='bold')
        for label in axes[i].get_xticklabels() + axes[i].get_yticklabels():
            label.set_fontname(times_new_roman)
            label.set_fontweight('bold')
        for spine in axes[i].spines.values():
            spine.set_linewidth(2)
        axes[i].grid(True, linestyle='--', linewidth=1.2, alpha=0.7)
    plt.tight_layout(pad=3.0)
    save_figure('data_comparison')
    plt.close()

def plot_distribution(original, cleaned, parameter):
    plt.figure(figsize=(12, 8))
    plt.hist(original, bins=30, color='royalblue', density=True, alpha=0.65, edgecolor='black', linewidth=1.5, label='Original Data')
    plt.hist(cleaned, bins=30, color='crimson', density=True, alpha=0.65, edgecolor='black', linewidth=1.5, label='Cleaned Data')
    ax = plt.gca()
    ax.tick_params(axis='both', which='major', direction='in', width=2.5, length=8, labelsize=20, pad=8)
    ax.tick_params(axis='both', which='minor', direction='in', width=1.5, length=4)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontweight('bold')
    for spine in ax.spines.values():
        spine.set_linewidth(2.0)
    x = np.linspace(original.min(), original.max(), 300)
    plt.plot(x, norm.pdf(x, *norm.fit(original.dropna())), 'b--', linewidth=3)
    plt.plot(x, norm.pdf(x, *norm.fit(cleaned.dropna())), 'r-', linewidth=3)
    plt.title(f' {parameter}', fontsize=24, fontweight='bold', pad=15)
    plt.ylabel("Probability Density", fontsize=20, fontweight='bold')
    plt.legend(edgecolor='black', framealpha=0.9, facecolor='white', loc='best', prop={'size': 16})
    ax.grid(True, linestyle=':', linewidth=1.5, alpha=0.6)
    plt.tight_layout(pad=3.0)
    save_figure(parameter + '_distribution')
    plt.close()

plot_boxplot(df_clean[cols])
plot_comparison(df_eng, df_clean, cols)
for col in cols:
    x_min, x_max = df_eng[col].min(), df_eng[col].max()
    original_norm = (df_eng[col] - x_min) / (x_max - x_min)
    cleaned_norm = (df_clean[col] - x_min) / (x_max - x_min)
    plot_distribution(original_norm, cleaned_norm, col)

# ============================== 指标输出 ==============================
print("\n=== Smoothness Comparison ===")
print(pd.DataFrame({
    'Original Smoothness': df_eng[cols].diff().abs().mean(),
    'Cleaned Smoothness': df_clean[cols].diff().abs().mean()
}))

print("\n=== Impact on Mean Values ===")
print(((df_clean[cols].mean() - df_eng[cols].mean()) / df_eng[cols].mean()).abs())

# ============================== 3D 散点图 ==============================
fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')
ax.scatter(df_clean[cols[0]], df_clean[cols[1]], df_clean[cols[2]], c=~outliers, cmap='coolwarm')
ax.set_xlabel(cols[0], fontproperties=label_font)
ax.set_ylabel(cols[1], fontproperties=label_font)
ax.set_zlabel(cols[2], fontproperties=label_font)
for label in ax.get_xticklabels() + ax.get_yticklabels() + ax.get_zticklabels():
    label.set_fontname(times_new_roman)
    label.set_fontweight('bold')
save_figure('3d_scatter')
plt.close()
